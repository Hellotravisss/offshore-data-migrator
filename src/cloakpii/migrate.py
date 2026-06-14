"""
Migration workflow: desensitize → encrypt → verify.

Supports:
  - 8 file formats: CSV, JSON, TXT, Excel, Parquet, XML, TSV, SQLite
  - Parallel batch processing (ThreadPoolExecutor)
  - Progress bar (tqdm)
  - Compliance validation (GDPR, PDPA, CCPA, LGPD, PIPL)
  - Integrity manifests (SHA-256)
  - Audit trail logging
  - Compression (gzip)
  - Resume (skip already processed files)
"""

from __future__ import annotations

import gzip
import json
import logging
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from .crypto import encrypt_file_with_key, derive_key, CryptoError, SALT_LEN
from .pii import (
    desensitize_csv,
    desensitize_json,
    desensitize_text,
    desensitize_excel,
    desensitize_parquet,
    desensitize_xml,
    desensitize_tsv,
    desensitize_sqlite,
)

logger = logging.getLogger("CloakPII")

# Threshold for large file warning (100 MB)
LARGE_FILE_THRESHOLD = 100 * 1024 * 1024


@dataclass
class MigrationReport:
    """Summary of a migration run."""
    started_at: str = ""
    finished_at: str = ""
    dry_run: bool = False
    target: str = ""
    files_processed: list = field(default_factory=list)
    files_encrypted: list = field(default_factory=list)
    files_skipped: list = field(default_factory=list)
    pii_reports: dict = field(default_factory=dict)
    errors: list = field(default_factory=list)
    workers: int = 1
    compliance_profile: str = ""
    compliance_violations: list = field(default_factory=list)
    manifest_hash: str = ""
    total_pii_masked: int = 0
    total_bytes_processed: int = 0

    def to_dict(self) -> dict:
        return {
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "dry_run": self.dry_run,
            "target": self.target,
            "files_processed": self.files_processed,
            "files_encrypted": self.files_encrypted,
            "files_skipped": self.files_skipped,
            "pii_reports": self.pii_reports,
            "errors": self.errors,
            "workers": self.workers,
            "compliance_profile": self.compliance_profile,
            "compliance_violations": self.compliance_violations,
            "manifest_hash": self.manifest_hash,
            "total_pii_masked": self.total_pii_masked,
            "total_bytes_processed": self.total_bytes_processed,
        }


def _classify_file(path: Path) -> Optional[str]:
    """Return file type for processing, or None if unsupported."""
    ext = path.suffix.lower()
    type_map = {
        ".csv": "csv",
        ".json": "json",
        ".xlsx": "excel",
        ".xls": "excel",
        ".parquet": "parquet",
        ".xml": "xml",
        ".tsv": "tsv",
        ".db": "sqlite",
        ".sqlite": "sqlite",
        ".sqlite3": "sqlite",
        ".txt": "text",
        ".log": "text",
        ".md": "text",
    }
    return type_map.get(ext)


SUPPORTED_EXTENSIONS = {
    ".csv", ".json", ".txt", ".log", ".md",
    ".xlsx", ".xls", ".parquet",
    ".xml", ".tsv",
    ".db", ".sqlite", ".sqlite3",
}


def run_migration(
    source_dir: Path,
    output_dir: Path,
    password: str,
    target: str = "singapore",
    dry_run: bool = False,
    workers: int = 1,
    batch_size: int = 0,
    show_progress: bool = True,
    compliance_profile: str = "",
    generate_manifest: bool = True,
    audit_log_path: Optional[Path] = None,
    compress: bool = False,
    skip_patterns: Optional[list[str]] = None,
    resume: bool = False,
    state=None,  # MigrationState for incremental processing
    mode: str = "mask",  # "mask" (irreversible) or "tokenize" (reversible pseudonyms)
) -> MigrationReport:
    """
    Execute migration pipeline: discover → desensitize → encrypt → verify.

    Args:
        source_dir: Directory containing files to migrate.
        output_dir: Directory for desensitized + encrypted output.
        password: Encryption password.
        target: Target jurisdiction name.
        dry_run: If True, only report what would happen.
        workers: Number of parallel workers (1 = sequential).
        batch_size: Max files to process (0 = all).
        show_progress: Show tqdm progress bar.
        compliance_profile: Name of compliance profile to validate against.
        generate_manifest: Generate SHA-256 manifest after migration.
        audit_log_path: Path for audit log file (JSON Lines).
        compress: Compress encrypted output with gzip.
        skip_patterns: Glob patterns for files to skip.
        resume: Skip files already processed (matched by path + SHA-256 hash);
            changed files are re-processed. Uses/creates a SQLite state DB.

    Returns:
        MigrationReport with full details.
    """
    report = MigrationReport(
        started_at=datetime.now(timezone.utc).isoformat(),
        dry_run=dry_run,
        target=target,
        workers=workers,
        compliance_profile=compliance_profile,
    )

    source_dir = Path(source_dir)
    output_dir = Path(output_dir)

    if not source_dir.exists():
        report.errors.append(f"Source directory not found: {source_dir}")
        logger.error(report.errors[-1])
        return report

    # Resume relies on a SQLite state DB (path + SHA-256 hash). If the caller
    # enabled resume via the Python API without supplying one, create the
    # default state DB inside the output directory (matches CLI behaviour).
    if resume and not dry_run and state is None:
        from .state import MigrationState
        state = MigrationState(output_dir / ".migration_state.db")

    # Derive the encryption key ONCE per run. PBKDF2 (480k iterations) is
    # intentionally expensive; running it per file made many-file migrations
    # pay that cost N times. A single run-level salt is stored in every file
    # header, so each output stays independently decryptable by password.
    enc_key = enc_salt = None
    if not dry_run:
        try:
            enc_salt = os.urandom(SALT_LEN)
            enc_key = derive_key(password, enc_salt)
        except CryptoError as exc:
            report.errors.append(str(exc))
            logger.error(str(exc))
            return report

    # Tokenization mode uses a deterministic, reversible pseudonymizer keyed by
    # the password (built once and shared across files).
    tokenizer = None
    if mode == "tokenize":
        from .tokenize import Tokenizer
        try:
            tokenizer = Tokenizer(password)
        except CryptoError as exc:
            report.errors.append(str(exc))
            logger.error(str(exc))
            return report

    # Initialize audit log
    audit = None
    if audit_log_path:
        from .audit import AuditLog
        audit = AuditLog(audit_log_path)
        audit.log_migration_start({
            "source": str(source_dir),
            "output": str(output_dir),
            "target": target,
            "dry_run": dry_run,
            "workers": workers,
            "compliance_profile": compliance_profile,
        })

    # Collect processable files
    files = []
    for f in sorted(source_dir.rglob("*")):
        if not f.is_file():
            continue
        if _classify_file(f) is None:
            continue
        # Skip patterns
        if skip_patterns:
            import fnmatch
            rel = str(f.relative_to(source_dir))
            if any(fnmatch.fnmatch(rel, pat) for pat in skip_patterns):
                continue
        # Resume: skip a file whose encrypted output already exists, unless the
        # state DB shows its content changed since it was processed (different
        # SHA-256 hash) — in which case it is re-processed rather than skipped.
        if resume and not dry_run:
            rel = f.relative_to(source_dir)
            enc_path = output_dir / "encrypted" / (str(rel) + ".enc")
            output_exists = enc_path.exists() or Path(str(enc_path) + ".gz").exists()
            if output_exists:
                recorded = state.get_recorded_hash(rel) if state is not None else None
                changed = recorded is not None and recorded != _get_file_hash(f)
                if not changed:
                    report.files_skipped.append(str(rel))
                    continue
        files.append(f)

    if not files:
        logger.warning(f"No processable files found (supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}).")
        report.finished_at = datetime.now(timezone.utc).isoformat()
        return report

    # Apply batch_size limit
    if batch_size > 0:
        files = files[:batch_size]
        logger.info(f"Batch mode: processing first {batch_size} of discovered files.")

    mode_label = "[DRY RUN]" if dry_run else "[MIGRATE]"

    # Warn about large files
    large_files = [(f, f.stat().st_size) for f in files if f.stat().st_size > LARGE_FILE_THRESHOLD]
    if large_files:
        total_large_mb = sum(s for _, s in large_files) / (1024 * 1024)
        logger.warning(
            f"{mode_label} Found {len(large_files)} large file(s) (>{LARGE_FILE_THRESHOLD // (1024*1024)}MB), "
            f"totaling {total_large_mb:.1f} MB. Processing may be slow."
        )
        for f, size in large_files[:5]:
            logger.warning(f"  {f.relative_to(source_dir)}: {size / (1024*1024):.1f} MB")

    if workers <= 1:
        _process_sequential(files, source_dir, output_dir, enc_key, enc_salt, dry_run, mode_label, report, show_progress, compress, audit, state, mode, tokenizer)
    else:
        _process_parallel(files, source_dir, output_dir, enc_key, enc_salt, dry_run, mode_label, report, workers, show_progress, compress, state, mode, tokenizer)

    # Compliance validation
    if compliance_profile and not dry_run:
        try:
            from .compliance import get_profile, validate_migration
            profile = get_profile(compliance_profile)
            violations = validate_migration(report.pii_reports, profile)
            report.compliance_violations = violations
            if violations:
                logger.warning(f"[COMPLIANCE] {len(violations)} violation(s) detected for {profile.name.upper()}")
                for v in violations:
                    logger.warning(f"  {v}")
            else:
                logger.info(f"[COMPLIANCE] ✓ Fully compliant with {profile.name.upper()}")
        except Exception as exc:
            logger.warning(f"[COMPLIANCE] Could not validate: {exc}")

    # Integrity manifest
    if generate_manifest and not dry_run and (output_dir / "encrypted").exists():
        try:
            from .integrity import write_manifest, compute_directory_hash
            manifest_path = output_dir / "manifest.json"
            write_manifest(output_dir / "encrypted", manifest_path)
            report.manifest_hash = compute_directory_hash(output_dir / "encrypted")
            logger.info(f"[INTEGRITY] Manifest written → {manifest_path}")
            logger.info(f"[INTEGRITY] Directory hash: {report.manifest_hash[:32]}...")
        except Exception as exc:
            logger.warning(f"[INTEGRITY] Could not generate manifest: {exc}")

    # Compute totals
    for pii_info in report.pii_reports.values():
        if isinstance(pii_info, dict):
            report.total_pii_masked += pii_info.get("values_masked", 0)
    for f in files:
        report.total_bytes_processed += f.stat().st_size

    report.finished_at = datetime.now(timezone.utc).isoformat()

    # Audit log end
    if audit:
        audit.log_migration_end(report.to_dict())

    logger.info(
        f"{mode_label} Complete. "
        f"Processed={len(report.files_processed)}, "
        f"Encrypted={len(report.files_encrypted)}, "
        f"Skipped={len(report.files_skipped)}, "
        f"PII masked={report.total_pii_masked}, "
        f"Errors={len(report.errors)}"
    )
    return report


def _process_sequential(files, source_dir, output_dir, enc_key, enc_salt, dry_run, mode_label, report, show_progress, compress, audit, state=None, transform_mode="mask", tokenizer=None):
    """Process files one by one."""
    iterator = _maybe_tqdm(files, "Processing", show_progress)
    for filepath in iterator:
        _process_single_file(filepath, source_dir, output_dir, enc_key, enc_salt, dry_run, mode_label, report, compress, audit, state, transform_mode, tokenizer)


def _process_parallel(files, source_dir, output_dir, enc_key, enc_salt, dry_run, mode_label, report, workers, show_progress, compress, state=None, transform_mode="mask", tokenizer=None):
    """Process files in parallel using ThreadPoolExecutor."""
    results = []
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_to_file = {
            executor.submit(
                _process_single_file_standalone,
                filepath, source_dir, output_dir, enc_key, enc_salt, dry_run, compress, state, transform_mode, tokenizer
            ): filepath
            for filepath in files
        }

        iterator = as_completed(future_to_file)
        if show_progress:
            try:
                from tqdm import tqdm
                iterator = tqdm(iterator, total=len(future_to_file), desc="Processing", unit="file")
            except ImportError:
                pass

        for future in iterator:
            filepath = future_to_file[future]
            try:
                result = future.result()
                results.append(result)
            except Exception as exc:
                rel = filepath.relative_to(source_dir)
                err = f"Failed processing {rel}: {exc}"
                report.errors.append(err)
                logger.error(err)

    for result in results:
        rel_str, file_type, pii_info, encrypted_rel, error = result
        report.files_processed.append(rel_str)
        if pii_info:
            report.pii_reports[rel_str] = pii_info
        if encrypted_rel:
            report.files_encrypted.append(encrypted_rel)
        if error:
            report.errors.append(error)


def _process_single_file(filepath, source_dir, output_dir, enc_key, enc_salt, dry_run, mode_label, report, compress, audit, state=None, transform_mode="mask", tokenizer=None):
    """Process a single file (sequential mode)."""
    rel = filepath.relative_to(source_dir)
    file_type = _classify_file(filepath)
    if file_type is None:
        return
    
    # Incremental migration check
    if state is not None:
        file_hash = _get_file_hash(filepath)
        if state.is_processed(rel, file_hash):
            logger.info(f"Skipping already processed file: {rel}")
            return
    report.files_processed.append(str(rel))

    # Warn about large files
    file_size = filepath.stat().st_size
    if file_size > LARGE_FILE_THRESHOLD:
        size_mb = file_size / (1024 * 1024)
        logger.warning(
            f"{mode_label} Large file: {rel} ({size_mb:.1f} MB). "
            f"Processing may be slow or use significant memory."
        )

    if dry_run:
        preview_info = _preview_file(filepath, file_type)
        report.pii_reports[str(rel)] = preview_info
        logger.info(f"{mode_label} Would process: {rel} (type={file_type}, pii_hits={preview_info.get('values_masked', 0)})")
        report.files_encrypted.append(str(rel) + " (planned)")
    else:
        try:
            desensitized_path = output_dir / "desensitized" / rel
            encrypted_path = output_dir / "encrypted" / (str(rel) + ".enc")

            pii_info = _desensitize_file(filepath, desensitized_path, file_type, transform_mode, tokenizer)

            # Mark as processed in state
            if state is not None:
                file_hash = _get_file_hash(filepath)
                pii_count = pii_info.get("values_masked", 0) if pii_info else 0
                state.mark_processed(rel, file_hash, pii_count)
            report.pii_reports[str(rel)] = pii_info
            logger.info(f"{mode_label} Desensitized: {rel} → {desensitized_path}")

            encrypt_file_with_key(desensitized_path, encrypted_path, enc_key, enc_salt)

            # Optional compression
            if compress:
                gz_path = Path(str(encrypted_path) + ".gz")
                with open(encrypted_path, "rb") as f_in:
                    with gzip.open(gz_path, "wb") as f_out:
                        f_out.writelines(f_in)
                encrypted_path.unlink()
                encrypted_path = gz_path

            report.files_encrypted.append(str(rel))
            logger.info(f"{mode_label} Encrypted: {rel} → {encrypted_path}")

            if audit:
                audit.log_file_processed(str(rel), pii_info.get("values_masked", 0))

        except Exception as exc:
            err = f"Failed processing {rel}: {exc}"
            report.errors.append(err)
            logger.error(err)
            if audit:
                audit.log_error(str(rel), str(exc))


def _process_single_file_standalone(filepath, source_dir, output_dir, enc_key, enc_salt, dry_run, compress, state=None, transform_mode="mask", tokenizer=None):
    """Process a single file (parallel mode — no shared state)."""
    rel = filepath.relative_to(source_dir)
    file_type = _classify_file(filepath)
    if file_type is None:
        return
    
    # Incremental migration check
    if state is not None:
        file_hash = _get_file_hash(filepath)
        if state.is_processed(rel, file_hash):
            logger.info(f"Skipping already processed file: {rel}")
            return (str(rel), None, None, None, None)

    if dry_run:
        preview_info = _preview_file(filepath, file_type)
        return (str(rel), file_type, preview_info, str(rel) + " (planned)", None)
    else:
        try:
            desensitized_path = output_dir / "desensitized" / rel
            encrypted_path = output_dir / "encrypted" / (str(rel) + ".enc")

            pii_info = _desensitize_file(filepath, desensitized_path, file_type, transform_mode, tokenizer)

            # Mark as processed in state
            if state is not None:
                file_hash = _get_file_hash(filepath)
                pii_count = pii_info.get("values_masked", 0) if pii_info else 0
                state.mark_processed(rel, file_hash, pii_count)
            encrypt_file_with_key(desensitized_path, encrypted_path, enc_key, enc_salt)

            if compress:
                gz_path = Path(str(encrypted_path) + ".gz")
                with open(encrypted_path, "rb") as f_in:
                    with gzip.open(gz_path, "wb") as f_out:
                        f_out.writelines(f_in)
                encrypted_path.unlink()
                encrypted_path = gz_path

            return (str(rel), file_type, pii_info, str(rel), None)
        except Exception as exc:
            logger.error(f"Failed processing {rel}: {exc}")
            return (str(rel), file_type, None, None, f"Failed processing {rel}: {exc}")


def _maybe_tqdm(iterable, desc, show_progress):
    """Wrap iterable with tqdm if available and requested."""
    if not show_progress:
        return iterable
    try:
        from tqdm import tqdm
        return tqdm(iterable, desc=desc, unit="file")
    except ImportError:
        return iterable


def _preview_file(filepath: Path, file_type: str) -> dict:
    """Count PII occurrences without modifying the file."""
    info = {"fields_masked": [], "values_masked": 0, "rows_processed": 0}
    try:
        if file_type == "csv":
            import csv
            from .pii import _is_pii_field, mask_value, mask_generic
            with open(filepath, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames or []
                info["fields_masked"] = [fn for fn in fieldnames if _is_pii_field(fn)]
                for row in reader:
                    info["rows_processed"] += 1
                    for fn in fieldnames:
                        val = row.get(fn, "")
                        masked = mask_value(val)
                        if masked == val and _is_pii_field(fn) and val.strip():
                            masked = mask_generic(val)
                        if masked != val:
                            info["values_masked"] += 1
        elif file_type == "json":
            from .pii import _desensitize_json_node, DesensitizeReport
            with open(filepath, encoding="utf-8") as f:
                data = json.load(f)
            r = DesensitizeReport()
            _desensitize_json_node(data, r)
            info["fields_masked"] = r.fields_masked
            info["values_masked"] = r.values_masked
            info["rows_processed"] = len(data) if isinstance(data, list) else 1
        elif file_type == "excel":
            info = _preview_excel(filepath)
        elif file_type == "parquet":
            info = _preview_parquet(filepath)
        elif file_type == "xml":
            info = _preview_xml(filepath)
        elif file_type == "tsv":
            info = _preview_tsv(filepath)
        elif file_type == "sqlite":
            info = _preview_sqlite(filepath)
        else:  # text
            text = filepath.read_text(encoding="utf-8")
            from .pii import desensitize_text
            _, count = desensitize_text(text)
            info["values_masked"] = count
            info["rows_processed"] = 1
    except Exception as exc:
        info["error"] = str(exc)
    return info


def _preview_excel(filepath: Path) -> dict:
    import openpyxl
    from .pii import _is_pii_field, mask_value, mask_generic
    info = {"fields_masked": [], "values_masked": 0, "rows_processed": 0}
    wb = openpyxl.load_workbook(filepath, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        fieldnames = [str(v) if v else "" for v in rows[0]]
        for fn in fieldnames:
            if _is_pii_field(fn) and fn not in info["fields_masked"]:
                info["fields_masked"].append(fn)
        for row in rows[1:]:
            info["rows_processed"] += 1
            for idx, val in enumerate(row):
                if val is None:
                    continue
                original = str(val)
                fn = fieldnames[idx] if idx < len(fieldnames) else ""
                masked = mask_value(original)
                if masked == original and _is_pii_field(fn) and original.strip():
                    masked = mask_generic(original)
                if masked != original:
                    info["values_masked"] += 1
    wb.close()
    return info


def _preview_parquet(filepath: Path) -> dict:
    import pyarrow.parquet as pq
    import pyarrow as pa
    from .pii import _is_pii_field, mask_value, mask_generic
    info = {"fields_masked": [], "values_masked": 0, "rows_processed": 0}
    table = pq.read_table(filepath)
    info["rows_processed"] = table.num_rows
    for i, f in enumerate(table.schema):
        if not (pa.types.is_string(f.type) or pa.types.is_large_string(f.type)):
            continue
        if _is_pii_field(f.name):
            info["fields_masked"].append(f.name)
        for val in table.column(i).to_pylist():
            if val is None:
                continue
            original = str(val)
            masked = mask_value(original)
            if masked == original and _is_pii_field(f.name) and original.strip():
                masked = mask_generic(original)
            if masked != original:
                info["values_masked"] += 1
    return info


def _preview_xml(filepath: Path) -> dict:
    import xml.etree.ElementTree as ET
    from .pii import _is_pii_field, mask_value, mask_generic
    info = {"fields_masked": [], "values_masked": 0, "rows_processed": 0}
    tree = ET.parse(filepath)
    root = tree.getroot()

    def _count(elem):
        info["rows_processed"] += 1
        if elem.text and elem.text.strip():
            original = elem.text
            masked = mask_value(original)
            if masked == original and _is_pii_field(elem.tag) and original.strip():
                masked = mask_generic(original)
            if masked != original:
                info["values_masked"] += 1
                if elem.tag not in info["fields_masked"]:
                    info["fields_masked"].append(elem.tag)
        for attr_name, attr_val in elem.attrib.items():
            masked = mask_value(attr_val)
            if masked == attr_val and _is_pii_field(attr_name) and attr_val.strip():
                masked = mask_generic(attr_val)
            if masked != attr_val:
                info["values_masked"] += 1
        for child in elem:
            _count(child)

    _count(root)
    return info


def _preview_tsv(filepath: Path) -> dict:
    import csv
    from .pii import _is_pii_field, mask_value, mask_generic
    info = {"fields_masked": [], "values_masked": 0, "rows_processed": 0}
    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        fieldnames = reader.fieldnames or []
        info["fields_masked"] = [fn for fn in fieldnames if _is_pii_field(fn)]
        for row in reader:
            info["rows_processed"] += 1
            for fn in fieldnames:
                val = row.get(fn, "")
                masked = mask_value(val)
                if masked == val and _is_pii_field(fn) and val.strip():
                    masked = mask_generic(val)
                if masked != val:
                    info["values_masked"] += 1
    return info


def _preview_sqlite(filepath: Path) -> dict:
    import sqlite3
    from .pii import _is_pii_field, mask_value, mask_generic
    info = {"fields_masked": [], "values_masked": 0, "rows_processed": 0}
    conn = sqlite3.connect(filepath)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = [row[0] for row in cursor.fetchall()]
    for table_name in tables:
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns = cursor.fetchall()
        string_cols = [col[1] for col in columns
                       if "TEXT" in (col[2] or "").upper() or "CHAR" in (col[2] or "").upper() or col[2] == ""]
        col_names = [col[1] for col in columns]
        for cn in string_cols:
            if _is_pii_field(cn) and cn not in info["fields_masked"]:
                info["fields_masked"].append(cn)
        cursor.execute(f"SELECT * FROM {table_name};")
        for row in cursor.fetchall():
            info["rows_processed"] += 1
            for idx, cn in enumerate(col_names):
                if cn not in string_cols or row[idx] is None:
                    continue
                original = str(row[idx])
                masked = mask_value(original)
                if masked == original and _is_pii_field(cn) and original.strip():
                    masked = mask_generic(original)
                if masked != original:
                    info["values_masked"] += 1
    conn.close()
    return info


def _desensitize_file(filepath: Path, output_path: Path, file_type: str,
                      mode="mask", tokenizer=None) -> dict:
    """Transform a file (mask/tokenize/detokenize) and return PII report info."""
    desensitizers = {
        "csv": desensitize_csv,
        "json": desensitize_json,
        "excel": desensitize_excel,
        "parquet": desensitize_parquet,
        "xml": desensitize_xml,
        "tsv": desensitize_tsv,
        "sqlite": desensitize_sqlite,
    }

    if file_type in desensitizers:
        r = desensitizers[file_type](filepath, output_path, mode=mode, tokenizer=tokenizer)
    elif file_type == "text":
        text = filepath.read_text(encoding="utf-8")
        masked, count = desensitize_text(text, mode=mode, tokenizer=tokenizer)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(masked, encoding="utf-8")
        return {"fields_masked": [], "values_masked": count, "rows_processed": 1}
    else:
        raise ValueError(f"Unsupported file type: {file_type}")

    return {
        "fields_masked": r.fields_masked,
        "values_masked": r.values_masked,
        "rows_processed": r.rows_processed,
    }

def _get_file_hash(path: Path) -> str:
    """Calculate SHA256 hash of a file for state tracking."""
    import hashlib
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def decrypt_tree(input_dir: Path, output_dir: Path, password: str) -> dict:
    """Decrypt a whole migration output tree back to plaintext.

    Walks ``input_dir`` for ``*.enc`` / ``*.enc.gz`` files (typically the
    ``encrypted/`` directory produced by a migration), decrypts each, and
    recreates the relative directory structure under ``output_dir`` with the
    ``.enc``/``.gz`` suffixes stripped.

    The key is derived once per distinct salt (a single migration run shares
    one salt across all files), so PBKDF2 runs once rather than per file.
    Returns a summary dict with decrypted/failed counts.
    """
    from .crypto import decrypt_data_with_key, read_salt, derive_key, CryptoError

    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    key_cache: dict[bytes, bytes] = {}
    decrypted: list[str] = []
    errors: list[str] = []

    for enc_path in sorted(input_dir.rglob("*")):
        if not enc_path.is_file():
            continue
        name = enc_path.name
        if name.endswith(".enc.gz"):
            blob = gzip.decompress(enc_path.read_bytes())
            rel = enc_path.relative_to(input_dir).with_name(name[: -len(".enc.gz")])
        elif name.endswith(".enc"):
            blob = enc_path.read_bytes()
            rel = enc_path.relative_to(input_dir).with_name(name[: -len(".enc")])
        else:
            continue

        try:
            salt = read_salt(blob)
            if salt not in key_cache:
                key_cache[salt] = derive_key(password, salt)
            plaintext = decrypt_data_with_key(blob, key_cache[salt])
        except CryptoError as exc:
            errors.append(f"{enc_path.relative_to(input_dir)}: {exc}")
            logger.error(f"[DECRYPT] Failed: {enc_path} ({exc})")
            continue

        dest = output_dir / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(plaintext)
        decrypted.append(str(rel))
        logger.info(f"[DECRYPT] {enc_path.relative_to(input_dir)} → {dest}")

    return {
        "decrypted": decrypted,
        "errors": errors,
        "total_decrypted": len(decrypted),
        "total_errors": len(errors),
    }


def detokenize_tree(input_dir: Path, output_dir: Path, password: str) -> dict:
    """Reverse tokenization across a (decrypted) tree produced by ``migrate
    --mode tokenize``.

    Re-parses each supported file and replaces every ``tkz_...`` token with its
    original value (so Parquet/Excel/SQLite are handled correctly, not just
    text formats). Files of unknown type are copied through unchanged.
    Requires the same password used at tokenization time.
    """
    import shutil
    from .tokenize import Tokenizer

    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    tokenizer = Tokenizer(password)
    detokenized: list[str] = []
    errors: list[str] = []

    for f in sorted(input_dir.rglob("*")):
        if not f.is_file():
            continue
        rel = f.relative_to(input_dir)
        dest = output_dir / rel
        file_type = _classify_file(f)
        try:
            if file_type is not None:
                _desensitize_file(f, dest, file_type, mode="detokenize", tokenizer=tokenizer)
                detokenized.append(str(rel))
                logger.info(f"[DETOKENIZE] {rel}")
            else:
                dest.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy(f, dest)
        except Exception as exc:
            errors.append(f"{rel}: {exc}")
            logger.error(f"[DETOKENIZE] Failed: {rel} ({exc})")

    return {
        "detokenized": detokenized,
        "errors": errors,
        "total_detokenized": len(detokenized),
        "total_errors": len(errors),
    }


# Enhanced error handling (v1.2)
class MigrationError(Exception):
    """Base exception for migration errors."""
    pass

class FileProcessingError(MigrationError):
    """Error processing a specific file."""
    def __init__(self, filepath: str, original_error: Exception):
        self.filepath = filepath
        self.original_error = original_error
        super().__init__(f"Failed to process {filepath}: {original_error}")
