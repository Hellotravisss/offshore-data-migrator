"""
PII detection and desensitization for multiple file formats.

Supported PII types:
  - email        user@example.com  →  u***@e******.com
  - phone        555-123-4567      →  555-***-****
  - national_id  123-45-6789 (SSN) →  ***-**-6789
  - credit_card  4111111111111111  →  4111****1111
  - ip_address   192.168.1.100     →  192.168.*.*
  - chinese_id   110101199001011234 →  1101***********234
  - passport     AB1234567         →  AB***4567
  - bank_account 1234567890123456  →  1234********3456
  - iban         GB29NWBK60161331926819 →  GB29****6819
  - mac_address  00:1B:44:11:3A:B7 →  00:1B:**:**:**:B7

Supported file formats:
  CSV, JSON, Excel (.xlsx), Parquet (.parquet), XML, TSV, SQLite

Field-name heuristics (for column names):
  If a column name contains keywords like 'name', 'email', 'phone', 'ssn',
  'address', 'id_number', the value is masked regardless of content pattern.
"""

import csv
import json
import re
import sqlite3
import shutil
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


# ---------------------------------------------------------------------------
# PII patterns
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
# Phone numbers must carry a signal that they ARE phone numbers — either an
# international "+CC" prefix or at least one separator between digit groups.
# This avoids masking bare integers (order IDs, counts, timestamps), which the
# old greedy pattern swept up as false positives.
PHONE_RE = re.compile(
    r"\+\d{1,3}[-.\s]?\d{2,4}[-.\s]?\d{3,4}[-.\s]?\d{3,4}"  # +CC international
    r"|"
    r"\(?\d{2,4}\)?[-.\s]\d{3,4}[-.\s]?\d{3,4}"            # domestic, ≥1 separator
)
SSN_RE = re.compile(r"\b\d{3}-\d{2}-\d{4}\b")
CREDIT_CARD_RE = re.compile(r"\b\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b")
IP_RE = re.compile(r"\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b")

# New PII patterns
CHINESE_ID_RE = re.compile(r"\b\d{17}[\dXx]\b")
PASSPORT_RE = re.compile(r"\b[A-Z]{1,2}\d{6,9}\b")
BANK_ACCOUNT_RE = re.compile(r"\b\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4,7}\b")
DATE_OF_BIRTH_RE = re.compile(
    r"\b(?:19|20)\d{2}[-/](?:0[1-9]|1[0-2])[-/](?:0[1-9]|[12]\d|3[01])\b|"
    r"\b(?:0[1-9]|1[0-2])[-/](?:0[1-9]|[12]\d|3[01])[-/](?:19|20)\d{2}\b"
)
IBAN_RE = re.compile(r"\b[A-Z]{2}\d{2}[A-Z0-9]{4}\d{7,16}\b")
MAC_ADDRESS_RE = re.compile(r"\b(?:[0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}\b")

# Column-name keywords that trigger name-based masking
NAME_KEYWORDS = {
    # English keywords
    "name", "email", "phone", "ssn", "address", "id_number",
    "national_id", "credit_card", "fullname", "full_name",
    "mobile", "tel", "telephone", "passport", "bank_account",
    "iban", "date_of_birth", "dob", "chinese_id", "mac_address",
    "license_number", "driver_license", "account_number",
    # Chinese keywords (重要：针对亚洲市场)
    "姓名", "邮箱", "电子邮箱", "手机", "手机号", "电话", "电话号码",
    "身份证", "身份证号", "身份证号码", "地址", "住址", "银行账户",
    "银行卡", "银行卡号", "护照", "护照号", "信用卡", "信用卡号",
    "出生日期", "生日", "mac地址", "驾照", "驾照号",
}


# ---------------------------------------------------------------------------
# Masking helpers
# ---------------------------------------------------------------------------

def mask_email(val: str) -> str:
    """user@example.com → u***@e******.com"""
    m = EMAIL_RE.search(val)
    if not m:
        return val
    local, domain = m.group().split("@", 1)
    masked_local = local[0] + "***" if len(local) > 1 else "***"
    dname, tld = domain.rsplit(".", 1) if "." in domain else (domain, "")
    masked_domain = dname[0] + "******" if dname else "******"
    result = f"{masked_local}@{masked_domain}.{tld}"
    return val[: m.start()] + result + val[m.end():]


def mask_phone(val: str) -> str:
    """Mask phone number, keeping country code (if any) and last 2 digits visible.
    
    Examples:
        555-123-4567 → 555-***-**67
        +1-555-123-4567 → +1-***-***-**67
        +86-138-1234-5678 → +86-***-****-**78
        416-555-0123 → 416-***-**23
    """
    digits = re.findall(r"\d", val)
    if len(digits) < 7:
        return val
    
    # Detect if there's a country code prefix (+ or 00)
    has_plus = val.strip().startswith("+")
    
    # Keep last 2 digits visible, mask the rest
    if has_plus:
        # For international: keep +CC and last 2
        # Find where country code ends (after first group of 1-3 digits following +)
        cc_match = re.match(r"(\+\d{1,3}[-.\s]?)", val)
        if cc_match:
            prefix = cc_match.group(1)
            rest = val[len(prefix):]
            # Mask all digits in rest except last 2
            result = []
            digit_count = len(re.findall(r"\d", rest))
            digits_to_mask = digit_count - 2
            masked_so_far = 0
            for ch in rest:
                if ch.isdigit():
                    if masked_so_far < digits_to_mask:
                        result.append("*")
                        masked_so_far += 1
                    else:
                        result.append(ch)
                else:
                    result.append(ch)
            return prefix + "".join(result)
    
    # For domestic: keep first 3 digits and last 2
    result = []
    digit_idx = 0
    total_digits = len(digits)
    for ch in val:
        if ch.isdigit():
            if digit_idx < 3 or digit_idx >= total_digits - 2:
                result.append(ch)
            else:
                result.append("*")
            digit_idx += 1
        else:
            result.append(ch)
    return "".join(result)


def mask_ssn(val: str) -> str:
    return "***-**-" + val[-4:]


def luhn_valid(number: str) -> bool:
    """Return True if the digit string passes the Luhn checksum (all card
    networks use it). Filters out random numeric strings that merely look
    card-shaped."""
    digits = [int(c) for c in number if c.isdigit()]
    if len(digits) < 12:
        return False
    checksum = 0
    parity = len(digits) % 2
    for i, d in enumerate(digits):
        if i % 2 == parity:
            d *= 2
            if d > 9:
                d -= 9
        checksum += d
    return checksum % 10 == 0


def mask_credit_card(val: str) -> str:
    clean = re.sub(r"[\s-]", "", val)
    if len(clean) < 8:
        return val
    # Only treat as a credit card if it passes the Luhn checksum; otherwise
    # leave it for a more specific matcher (e.g. bank account) to handle.
    if not luhn_valid(clean):
        return val
    return clean[:4] + "****" + clean[-4:]


def mask_ip(val: str) -> str:
    parts = val.split(".")
    if len(parts) != 4:
        return val
    return f"{parts[0]}.{parts[1]}.*.*"


def mask_generic(val: str) -> str:
    """Generic masking: keep first char, replace rest with ***."""
    if not val or len(val) < 2:
        return "***"
    return val[0] + "***"


def mask_chinese_id(val: str) -> str:
    """110101199001011234 → 1101***********234"""
    if len(val) < 8:
        return val
    return val[:4] + "*" * (len(val) - 7) + val[-3:]


def mask_passport(val: str) -> str:
    """AB1234567 → AB***4567"""
    if len(val) < 5:
        return val
    return val[:2] + "***" + val[-4:]


def mask_bank_account(val: str) -> str:
    """1234567890123456 → 1234********3456"""
    clean = re.sub(r"[\s-]", "", val)
    if len(clean) < 8:
        return val
    return clean[:4] + "*" * (len(clean) - 8) + clean[-4:]


def mask_dob(val: str) -> str:
    """1990-01-15 → ****-**-15"""
    # Keep last 2 digits (day) visible
    parts = re.split(r"[-/]", val)
    if len(parts) == 3:
        return "****-**-" + parts[-1]
    return val


def mask_iban(val: str) -> str:
    """GB29NWBK60161331926819 → GB29****6819"""
    if len(val) < 8:
        return val
    return val[:4] + "****" + val[-4:]


def mask_mac(val: str) -> str:
    """00:1B:44:11:3A:B7 → 00:1B:**:**:**:B7"""
    parts = re.split(r"[:-]", val)
    if len(parts) != 6:
        return val
    return f"{parts[0]}:{parts[1]}:**:**:**:{parts[5]}"


# ---------------------------------------------------------------------------
# Auto-detect & mask a single string value
# ---------------------------------------------------------------------------

def mask_value(val: str) -> str:
    """Apply all regex-based PII masks to a string value."""
    if not isinstance(val, str):
        return val
    result = val
    # Order matters: more specific patterns first
    result = SSN_RE.sub(lambda m: mask_ssn(m.group()), result)
    result = CHINESE_ID_RE.sub(lambda m: mask_chinese_id(m.group()), result)
    result = IBAN_RE.sub(lambda m: mask_iban(m.group()), result)
    result = CREDIT_CARD_RE.sub(lambda m: mask_credit_card(m.group()), result)
    result = BANK_ACCOUNT_RE.sub(lambda m: mask_bank_account(m.group()), result)
    result = PASSPORT_RE.sub(lambda m: mask_passport(m.group()), result)
    result = EMAIL_RE.sub(lambda m: mask_email(m.group()), result)
    result = IP_RE.sub(lambda m: mask_ip(m.group()), result)
    result = MAC_ADDRESS_RE.sub(lambda m: mask_mac(m.group()), result)
    result = PHONE_RE.sub(lambda m: mask_phone(m.group()), result)
    return result


def _is_pii_field(field_name: str) -> bool:
    """Check if a column/field name suggests PII content."""
    lower = field_name.lower().replace(" ", "_").replace("-", "_")
    return any(kw in lower for kw in NAME_KEYWORDS)


# ---------------------------------------------------------------------------
# Desensitization stats
# ---------------------------------------------------------------------------

@dataclass
class DesensitizeReport:
    """Tracks what was masked during a desensitization pass."""
    fields_masked: list = field(default_factory=list)
    values_masked: int = 0
    rows_processed: int = 0


# ---------------------------------------------------------------------------
# Text
# ---------------------------------------------------------------------------

def desensitize_text(text: str) -> tuple:
    """Mask PII in plain text. Returns (masked_text, count_of_replacements)."""
    count = 0
    
    # Define all pattern-masker pairs explicitly
    pattern_maskers = [
        (SSN_RE, mask_ssn),
        (CHINESE_ID_RE, mask_chinese_id),
        (IBAN_RE, mask_iban),
        (CREDIT_CARD_RE, mask_credit_card),
        (BANK_ACCOUNT_RE, mask_bank_account),
        (PASSPORT_RE, mask_passport),
        (EMAIL_RE, mask_email),
        (IP_RE, mask_ip),
        (MAC_ADDRESS_RE, mask_mac),
        (PHONE_RE, mask_phone),
    ]
    
    # Process each pattern
    for pattern, masker_fn in pattern_maskers:
        def make_replacer(fn):
            """Create a replacer function with proper closure."""
            def _replace(m):
                nonlocal count
                count += 1
                return fn(m.group())
            return _replace
        
        text = pattern.sub(make_replacer(masker_fn), text)

    # Apply custom patterns
    for name, pattern in CUSTOM_PII_PATTERNS:
        try:
            custom_re = re.compile(pattern)
            def make_custom_replacer(pname):
                def _replace(m):
                    nonlocal count
                    count += 1
                    return f"[{pname}]"
                return _replace
            text = custom_re.sub(make_custom_replacer(name), text)
        except re.error:
            continue

    return text, count


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def desensitize_csv(input_path: Path, output_path: Path) -> DesensitizeReport:
    """Read CSV, mask PII in every cell, write to output_path."""
    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    with open(input_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []

        # Determine which fields need name-based masking
        pii_fields = [fn for fn in fieldnames if _is_pii_field(fn)]
        report.fields_masked = list(pii_fields)

        rows = []
        for row in reader:
            report.rows_processed += 1
            for fn in fieldnames:
                original = row.get(fn, "")
                masked = mask_value(original)
                # If field name suggests PII and regex didn't catch it, do generic mask
                if masked == original and _is_pii_field(fn) and original.strip():
                    masked = mask_generic(original)
                if masked != original:
                    report.values_masked += 1
                row[fn] = masked
            rows.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return report


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------

def _desensitize_json_node(node: Any, report: DesensitizeReport, parent_key: str = "") -> Any:
    """Recursively walk a JSON structure and mask PII values."""
    if isinstance(node, dict):
        result = {}
        for k, v in node.items():
            result[k] = _desensitize_json_node(v, report, parent_key=k)
        return result
    elif isinstance(node, list):
        return [_desensitize_json_node(item, report, parent_key=parent_key) for item in node]
    elif isinstance(node, str):
        original = node
        masked = mask_value(node)
        if masked == original and _is_pii_field(parent_key) and original.strip():
            masked = mask_generic(original)
        if masked != original:
            report.values_masked += 1
            if parent_key and parent_key not in report.fields_masked:
                report.fields_masked.append(parent_key)
        return masked
    else:
        return node


def desensitize_json(input_path: Path, output_path: Path) -> DesensitizeReport:
    """Read JSON, mask PII recursively, write to output_path."""
    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        report.rows_processed = len(data)

    masked_data = _desensitize_json_node(data, report)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(masked_data, f, indent=2, ensure_ascii=False)

    return report


# ---------------------------------------------------------------------------
# Excel (.xlsx)
# ---------------------------------------------------------------------------

def desensitize_excel(input_path: Path, output_path: Path) -> DesensitizeReport:
    """Read Excel workbook, mask PII in all sheets, write to output_path."""
    import openpyxl

    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    wb = openpyxl.load_workbook(input_path, read_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            continue

        # First row is header
        header_cells = rows[0]
        fieldnames = [cell.value or "" for cell in header_cells]
        pii_fields = [fn for fn in fieldnames if _is_pii_field(str(fn))]
        for pf in pii_fields:
            if pf not in report.fields_masked:
                report.fields_masked.append(pf)

        for row in rows[1:]:
            report.rows_processed += 1
            for idx, cell in enumerate(row):
                if cell.value is None:
                    continue
                original = str(cell.value)
                fn = fieldnames[idx] if idx < len(fieldnames) else ""
                masked = mask_value(original)
                if masked == original and _is_pii_field(str(fn)) and original.strip():
                    masked = mask_generic(original)
                if masked != original:
                    report.values_masked += 1
                    cell.value = masked

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return report


# ---------------------------------------------------------------------------
# Parquet (.parquet)
# ---------------------------------------------------------------------------

def desensitize_parquet(input_path: Path, output_path: Path) -> DesensitizeReport:
    """Read Parquet file, mask PII in string columns, write to output_path."""
    import pyarrow.parquet as pq
    import pyarrow as pa

    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    table = pq.read_table(input_path)
    schema = table.schema

    # Identify string columns and PII field names
    pii_fields = []
    string_cols = []
    for i, fld in enumerate(schema):
        if pa.types.is_string(fld.type) or pa.types.is_large_string(fld.type):
            string_cols.append(i)
            if _is_pii_field(fld.name):
                pii_fields.append(fld.name)

    report.fields_masked = list(pii_fields)
    report.rows_processed = table.num_rows

    # Process each string column
    new_columns = []
    for i, fld in enumerate(schema):
        col = table.column(i)
        if i in string_cols:
            masked_values = []
            for val in col.to_pylist():
                if val is None:
                    masked_values.append(None)
                    continue
                original = str(val)
                masked = mask_value(original)
                if masked == original and _is_pii_field(fld.name) and original.strip():
                    masked = mask_generic(original)
                if masked != original:
                    report.values_masked += 1
                masked_values.append(masked)
            new_columns.append(pa.array(masked_values, type=fld.type))
        else:
            new_columns.append(col)

    new_table = pa.table(dict(zip([f.name for f in schema], new_columns)))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    pq.write_table(new_table, output_path)

    return report


# ---------------------------------------------------------------------------
# XML (.xml)
# ---------------------------------------------------------------------------

def desensitize_xml(input_path: Path, output_path: Path) -> DesensitizeReport:
    """Read XML file, mask PII in text content and attributes, write to output_path."""
    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    tree = ET.parse(input_path)
    root = tree.getroot()

    def _mask_element(elem):
        """Recursively mask PII in an XML element."""
        # Mask text content
        if elem.text and elem.text.strip():
            original = elem.text
            masked = mask_value(original)
            if masked != original:
                report.values_masked += 1
                elem.text = masked

        # Mask tail (text after closing tag)
        if elem.tail and elem.tail.strip():
            original = elem.tail
            masked = mask_value(original)
            if masked != original:
                report.values_masked += 1
                elem.tail = masked

        # Mask attributes
        for attr_name, attr_val in list(elem.attrib.items()):
            masked = mask_value(attr_val)
            # Also check if attribute name suggests PII
            if masked == attr_val and _is_pii_field(attr_name) and attr_val.strip():
                masked = mask_generic(attr_val)
            if masked != attr_val:
                report.values_masked += 1
                if attr_name not in report.fields_masked:
                    report.fields_masked.append(attr_name)
                elem.set(attr_name, masked)

        # Check if element tag name suggests PII
        if _is_pii_field(elem.tag) and elem.text and elem.text.strip():
            original = elem.text
            masked = mask_value(original)
            if masked == original:
                masked = mask_generic(original)
            if masked != original:
                report.values_masked += 1
                if elem.tag not in report.fields_masked:
                    report.fields_masked.append(elem.tag)
                elem.text = masked

        report.rows_processed += 1

        # Recurse into children
        for child in elem:
            _mask_element(child)

    _mask_element(root)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Add indentation for readable output (Python 3.9+)
    try:
        ET.indent(tree, space="  ")
    except AttributeError:
        pass  # ET.indent not available in Python < 3.9
    tree.write(output_path, encoding="unicode", xml_declaration=True)

    return report


# ---------------------------------------------------------------------------
# TSV (Tab-Separated Values)
# ---------------------------------------------------------------------------

def desensitize_tsv(input_path: Path, output_path: Path) -> DesensitizeReport:
    """Read TSV file, mask PII in every cell, write to output_path."""
    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    with open(input_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        fieldnames = reader.fieldnames or []

        # Determine which fields need name-based masking
        pii_fields = [fn for fn in fieldnames if _is_pii_field(fn)]
        report.fields_masked = list(pii_fields)

        rows = []
        for row in reader:
            report.rows_processed += 1
            for fn in fieldnames:
                original = row.get(fn, "")
                masked = mask_value(original)
                # If field name suggests PII and regex didn't catch it, do generic mask
                if masked == original and _is_pii_field(fn) and original.strip():
                    masked = mask_generic(original)
                if masked != original:
                    report.values_masked += 1
                row[fn] = masked
            rows.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)

    return report


# ---------------------------------------------------------------------------
# SQLite (.db, .sqlite, .sqlite3)
# ---------------------------------------------------------------------------

def desensitize_sqlite(input_path: Path, output_path: Path) -> DesensitizeReport:
    """Copy SQLite database and mask PII in all string columns."""
    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    # Copy the database first
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(input_path, output_path)

    # Connect to the copy and mask in place
    conn = sqlite3.connect(output_path)
    cursor = conn.cursor()

    # Get all tables
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = [row[0] for row in cursor.fetchall()]

    for table_name in tables:
        # Get column info
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns = cursor.fetchall()
        # columns: (cid, name, type, notnull, dflt_value, pk)

        # Identify string columns
        string_cols = []
        for col in columns:
            col_name = col[1]
            col_type = col[2].upper() if col[2] else ""
            if "TEXT" in col_type or "CHAR" in col_type or "CLOB" in col_type or col_type == "":
                string_cols.append(col_name)

        if not string_cols:
            continue

        # Get all rows
        cursor.execute(f"SELECT * FROM {table_name};")
        rows = cursor.fetchall()
        col_names = [col[1] for col in columns]

        report.rows_processed += len(rows)

        # Track PII fields
        for col_name in string_cols:
            if _is_pii_field(col_name) and col_name not in report.fields_masked:
                report.fields_masked.append(col_name)

        # Process each row
        for row_idx, row in enumerate(rows):
            updates = {}
            for col_idx, col_name in enumerate(col_names):
                if col_name not in string_cols:
                    continue

                original = row[col_idx]
                if original is None:
                    continue

                original_str = str(original)
                masked = mask_value(original_str)

                # Field name heuristic
                if masked == original_str and _is_pii_field(col_name) and original_str.strip():
                    masked = mask_generic(original_str)

                if masked != original_str:
                    report.values_masked += 1
                    updates[col_name] = masked

            # Update the row if any changes
            if updates:
                set_clause = ", ".join(f"{k} = ?" for k in updates.keys())
                values = list(updates.values())
                # Use rowid to identify the row (SQLite internal)
                cursor.execute(
                    f"UPDATE {table_name} SET {set_clause} WHERE rowid = ?",
                    values + [row_idx + 1]
                )

    conn.commit()
    conn.close()

    return report

# Custom PII pattern support (v1.2)
CUSTOM_PII_PATTERNS: list[tuple[str, str]] = []

def register_custom_pii_pattern(name: str, pattern: str):
    """Register a custom PII detection pattern."""
    import re
    try:
        re.compile(pattern)
        CUSTOM_PII_PATTERNS.append((name, pattern))
    except re.error:
        raise ValueError(f"Invalid regex pattern for {name}: {pattern}")

def _apply_custom_patterns(text: str) -> list[tuple[str, str]]:
    """Apply registered custom PII patterns to text."""
    import re
    matches = []
    for name, pattern in CUSTOM_PII_PATTERNS:
        try:
            for m in re.finditer(pattern, text):
                matches.append((name, m.group()))
        except re.error:
            continue
    return matches
