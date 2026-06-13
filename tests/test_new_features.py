"""Tests for Excel and Parquet desensitization + batch/parallel migration."""

import json
from pathlib import Path

import pytest

# Fixtures shared across tests
@pytest.fixture
def sample_excel(tmp_path):
    """Create a sample Excel file with PII."""
    import openpyxl
    p = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["name", "email", "phone", "ssn", "department"])
    ws.append(["Alice Johnson", "alice@example.com", "555-123-4567", "123-45-6789", "Engineering"])
    ws.append(["Bob Smith", "bob@company.io", "416-555-0123", "987-65-4321", "Marketing"])
    wb.save(p)
    wb.close()
    return p


@pytest.fixture
def sample_parquet(tmp_path):
    """Create a sample Parquet file with PII."""
    import pyarrow as pa
    import pyarrow.parquet as pq
    p = tmp_path / "test.parquet"
    data = {
        "id": [1, 2, 3],
        "full_name": ["Alice Johnson", "Bob Smith", "Carol W"],
        "email": ["alice@example.com", "bob@company.io", "carol@startup.net"],
        "ip_address": ["192.168.1.100", "10.0.0.55", "172.16.0.1"],
        "department": ["Engineering", "Marketing", "Finance"],
    }
    pq.write_table(pa.table(data), p)
    return p


@pytest.fixture
def mixed_source_dir(tmp_path, sample_excel, sample_parquet):
    """Create a source directory with CSV, JSON, Excel, Parquet, and text files."""
    src = tmp_path / "source"
    src.mkdir()

    # Copy Excel and Parquet
    import shutil
    shutil.copy(sample_excel, src / "data.xlsx")
    shutil.copy(sample_parquet, src / "data.parquet")

    # CSV
    (src / "data.csv").write_text(
        "name,email,score\nAlice,alice@test.com,90\nBob,bob@test.com,80\n"
    )

    # JSON
    (src / "data.json").write_text(json.dumps([
        {"name": "Alice", "email": "alice@test.com", "score": 90},
        {"name": "Bob", "email": "bob@test.com", "score": 80},
    ]))

    # Text
    (src / "notes.txt").write_text("Contact alice@test.com for details.")

    return src


# ---------------------------------------------------------------------------
# Excel desensitization tests
# ---------------------------------------------------------------------------

class TestDesensitizeExcel:
    def test_masks_pii_fields(self, sample_excel, tmp_path):
        from piiguard.pii import desensitize_excel
        out = tmp_path / "out.xlsx"
        report = desensitize_excel(sample_excel, out)

        assert out.exists()
        assert report.values_masked > 0
        assert report.rows_processed == 2
        assert len(report.fields_masked) > 0

    def test_output_is_valid_xlsx(self, sample_excel, tmp_path):
        import openpyxl
        from piiguard.pii import desensitize_excel
        out = tmp_path / "out.xlsx"
        desensitize_excel(sample_excel, out)

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # Header + 2 data rows
        assert len(rows) == 3
        # Check that email was masked
        email_col = rows[0].index("email")
        assert "***" in str(rows[1][email_col])
        wb.close()

    def test_empty_excel(self, tmp_path):
        import openpyxl
        from piiguard.pii import desensitize_excel
        p = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.save(p)
        wb.close()

        out = tmp_path / "out.xlsx"
        report = desensitize_excel(p, out)
        assert report.values_masked == 0
        assert report.rows_processed == 0


# ---------------------------------------------------------------------------
# Parquet desensitization tests
# ---------------------------------------------------------------------------

class TestDesensitizeParquet:
    def test_masks_string_columns(self, sample_parquet, tmp_path):
        from piiguard.pii import desensitize_parquet
        out = tmp_path / "out.parquet"
        report = desensitize_parquet(sample_parquet, out)

        assert out.exists()
        assert report.values_masked > 0
        assert report.rows_processed == 3

    def test_output_is_valid_parquet(self, sample_parquet, tmp_path):
        import pyarrow.parquet as pq
        from piiguard.pii import desensitize_parquet
        out = tmp_path / "out.parquet"
        desensitize_parquet(sample_parquet, out)

        table = pq.read_table(out)
        assert table.num_rows == 3
        # Check email column was masked
        emails = table.column("email").to_pylist()
        for email in emails:
            assert "***" in email

    def test_preserves_non_string_columns(self, sample_parquet, tmp_path):
        import pyarrow.parquet as pq
        from piiguard.pii import desensitize_parquet
        out = tmp_path / "out.parquet"
        desensitize_parquet(sample_parquet, out)

        table = pq.read_table(out)
        ids = table.column("id").to_pylist()
        assert ids == [1, 2, 3]

    def test_pii_fields_detected(self, sample_parquet, tmp_path):
        from piiguard.pii import desensitize_parquet
        out = tmp_path / "out.parquet"
        report = desensitize_parquet(sample_parquet, out)
        # full_name, email, ip_address should be detected
        assert "full_name" in report.fields_masked
        assert "email" in report.fields_masked
        assert "ip_address" in report.fields_masked


# ---------------------------------------------------------------------------
# File classification tests
# ---------------------------------------------------------------------------

class TestFileClassification:
    def test_classify_excel(self):
        from piiguard.migrate import _classify_file
        assert _classify_file(Path("data.xlsx")) == "excel"
        assert _classify_file(Path("data.xls")) == "excel"

    def test_classify_parquet(self):
        from piiguard.migrate import _classify_file
        assert _classify_file(Path("data.parquet")) == "parquet"

    def test_classify_existing_types(self):
        from piiguard.migrate import _classify_file
        assert _classify_file(Path("data.csv")) == "csv"
        assert _classify_file(Path("data.json")) == "json"
        assert _classify_file(Path("data.txt")) == "text"
        assert _classify_file(Path("data.png")) is None


# ---------------------------------------------------------------------------
# Migration with new file types
# ---------------------------------------------------------------------------

class TestMigrationNewTypes:
    def test_migration_includes_excel_parquet(self, mixed_source_dir, tmp_path):
        from piiguard.migrate import run_migration
        out = tmp_path / "output"
        report = run_migration(
            source_dir=mixed_source_dir,
            output_dir=out,
            password="testpw",
            dry_run=False,
            show_progress=False,
        )
        # Should have processed all 5 files
        assert len(report.files_processed) == 5
        assert "data.xlsx" in report.files_processed
        assert "data.parquet" in report.files_processed
        assert len(report.errors) == 0

    def test_dry_run_reports_excel_parquet(self, mixed_source_dir, tmp_path):
        from piiguard.migrate import run_migration
        out = tmp_path / "output"
        report = run_migration(
            source_dir=mixed_source_dir,
            output_dir=out,
            password="testpw",
            dry_run=True,
            show_progress=False,
        )
        assert len(report.files_processed) == 5
        # Dry run should not create output files
        assert not (out / "encrypted").exists()

    def test_encrypted_excel_is_decryptable(self, mixed_source_dir, tmp_path):
        from piiguard.migrate import run_migration
        from piiguard.crypto import decrypt_file
        out = tmp_path / "output"
        run_migration(
            source_dir=mixed_source_dir,
            output_dir=out,
            password="testpw",
            show_progress=False,
        )
        # Decrypt the Excel .enc file and verify it's valid
        enc_path = out / "encrypted" / "data.xlsx.enc"
        dec_path = tmp_path / "decrypted.xlsx"
        decrypt_file(enc_path, dec_path, "testpw")

        import openpyxl
        wb = openpyxl.load_workbook(dec_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 2
        wb.close()


# ---------------------------------------------------------------------------
# Batch processing tests
# ---------------------------------------------------------------------------

class TestBatchProcessing:
    def test_batch_size_limits_files(self, mixed_source_dir, tmp_path):
        from piiguard.migrate import run_migration
        out = tmp_path / "output"
        report = run_migration(
            source_dir=mixed_source_dir,
            output_dir=out,
            password="testpw",
            batch_size=2,
            show_progress=False,
        )
        assert len(report.files_processed) == 2

    def test_parallel_workers(self, mixed_source_dir, tmp_path):
        from piiguard.migrate import run_migration
        out = tmp_path / "output"
        report = run_migration(
            source_dir=mixed_source_dir,
            output_dir=out,
            password="testpw",
            workers=3,
            show_progress=False,
        )
        assert len(report.files_processed) == 5
        assert len(report.errors) == 0

    def test_parallel_dry_run(self, mixed_source_dir, tmp_path):
        from piiguard.migrate import run_migration
        out = tmp_path / "output"
        report = run_migration(
            source_dir=mixed_source_dir,
            output_dir=out,
            password="testpw",
            workers=2,
            dry_run=True,
            show_progress=False,
        )
        assert len(report.files_processed) == 5

    def test_report_includes_workers(self, mixed_source_dir, tmp_path):
        from piiguard.migrate import run_migration
        out = tmp_path / "output"
        report = run_migration(
            source_dir=mixed_source_dir,
            output_dir=out,
            password="testpw",
            workers=4,
            show_progress=False,
        )
        d = report.to_dict()
        assert d["workers"] == 4


# ---------------------------------------------------------------------------
# Preview tests for new types
# ---------------------------------------------------------------------------

class TestPreviewNewTypes:
    def test_preview_excel(self, sample_excel):
        from piiguard.migrate import _preview_file
        info = _preview_file(sample_excel, "excel")
        assert info["values_masked"] > 0
        assert info["rows_processed"] == 2

    def test_preview_parquet(self, sample_parquet):
        from piiguard.migrate import _preview_file
        info = _preview_file(sample_parquet, "parquet")
        assert info["values_masked"] > 0
        assert info["rows_processed"] == 3
